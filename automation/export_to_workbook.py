#!/usr/bin/env python3
"""Export SQLite data to the Excel workbook's Current List sheet.

Writes ALL columns (A through AG, skipping computed R/S/T) so the
workbook's pivot table and formulas have the full dataset.
"""

import logging
import sqlite3
import sys
import time

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

CURRENT_LIST_SHEET = "Current List"

# Column layout: (excel_column_letter, db_field_name, value_formatter)
# Covers A through AG, skipping R/S/T (computed by workbook formulas).
# Column AH and beyond are also computed or unused.
EXPORT_COLUMNS = [
    ("A",  "detailing_due_date",            "date"),
    ("B",  "dept_due_date_previous",        "str"),
    ("C",  "com_number",                    "str"),
    ("D",  "manufacturing_location",        "str"),
    ("E",  "detailer",                      "str"),
    ("F",  "job_name",                      "str"),
    ("G",  "top_level_number",              "str"),
    ("H",  "description",                   "str"),
    ("I",  "build_date",                    "date"),
    ("J",  "build_cycle",                   "int"),
    ("K",  "department_hours",              "float"),
    ("L",  "percent_complete",              "percent"),
    ("M",  "remaining_hours",               "float"),
    ("N",  "actual_hours",                  "float"),
    ("O",  "week_ending_friday",            "date"),
    ("P",  "notes",                         "str"),
    ("Q",  "late",                          "int"),
    # R, S, T are COMPUTED — workbook formulas, skip
    ("U",  "checking_status",               "str"),
    ("V",  "target_dept_hours",             "float"),
    ("W",  "iec_internal_hours",            "float"),
    ("X",  "unit_detailing_start_date",     "date"),
    ("Y",  "unit_moved_to_checking_date",   "date"),
    ("Z",  "unit_detailing_completion_date","date"),
    ("AA", "actual_hours_to_detail_unit",   "float"),
    ("AB", "hour_variance",                 "float"),
    ("AC", "remaining_demand",              "float"),
    ("AD", "same_as",                       "str"),
    ("AE", "dr_checks",                     "str"),
    ("AF", "dvl_checks",                    "str"),
    ("AG", "hours_checking",                "float"),
]


def _format_value(val, fmt: str):
    """Convert a Python value to the appropriate Excel cell value."""
    if val is None:
        return None
    if fmt == "date":
        if isinstance(val, str):
            # ISO date string — convert to datetime for Excel
            from datetime import datetime
            try:
                return datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                return val
        return val  # already a date/datetime object
    if fmt == "percent":
        # Store as 0.0–1.0 decimal (Excel displays as percentage)
        return val
    if fmt == "float":
        return float(val) if val is not None else None
    if fmt == "int":
        return int(val) if val is not None else None
    return str(val) if val else None


def export_to_workbook(db_path: str, excel_path: str) -> int:
    """Export all units from SQLite to the workbook's Current List sheet.

    Returns the number of rows exported.
    """
    # Read all columns from SQLite
    db_fields = [col[1] for col in EXPORT_COLUMNS]
    select_cols = ", ".join(db_fields)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(f"SELECT {select_cols} FROM units ORDER BY detailing_due_date")
    rows = cur.fetchall()
    conn.close()

    log.info(f"Read {len(rows)} rows from SQLite")

    # Open workbook
    wb = load_workbook(excel_path, data_only=False, keep_vba=True)

    # Find the Current List sheet
    if CURRENT_LIST_SHEET not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{CURRENT_LIST_SHEET}' not found. Available: {available}"
        )
    ws = wb[CURRENT_LIST_SHEET]

    # Read header row to map column letters to actual column indices
    # The header row (row 1) should have the column names
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip()] = cell.column

    # Clear old data (keep header row)
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    # Build a mapping: for each export column, find the target column index
    # by matching the db_field_name to the header row value
    col_index_map = []
    for col_letter, db_field, fmt in EXPORT_COLUMNS:
        # Try to find by header name first
        target_col = header_map.get(db_field)
        if target_col is None:
            # Fall back to column letter
            from openpyxl.utils import column_index_from_string
            target_col = column_index_from_string(col_letter)
        col_index_map.append((target_col, db_field, fmt))

    # Write data rows
    for row_idx, db_row in enumerate(rows, start=2):
        for col_idx, (target_col, db_field, fmt) in enumerate(col_index_map):
            raw_val = db_row[db_field]
            cell_value = _format_value(raw_val, fmt)
            ws.cell(row=row_idx, column=target_col, value=cell_value)

    wb.save(excel_path)
    log.info(f"Exported {len(rows)} rows to {excel_path} ({CURRENT_LIST_SHEET} sheet)")
    return len(rows)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Export SQLite to Excel Current List")
    parser.add_argument("--db", required=True, help="Path to SQLite database")
    parser.add_argument("--excel-path", required=True, help="Path to Excel workbook")
    args = parser.parse_args()

    t0 = time.perf_counter()
    count = export_to_workbook(args.db, args.excel_path)
    elapsed = time.perf_counter() - t0
    print(f"Exported {count} rows in {elapsed:.1f}s")


if __name__ == "__main__":
    main()
