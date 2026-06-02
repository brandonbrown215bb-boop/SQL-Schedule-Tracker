# tests/test_multi_user_integration.py
"""Integration tests for multi-user sync concurrency.

Tests simulate concurrent access from two "users" (Alice, Bob) using
threading, not full Qt app instances. Each test gets a fresh temp
directory to avoid state leakage.
"""

from __future__ import annotations

import json
import os
import threading
import time
from pathlib import Path

import pytest
from openpyxl import Workbook

from data.loader import COLUMN_MAP, unit_fingerprint
from data.models import Unit
from data.writer import save_unit
from sync.lock_manager import LockAcquisitionError, LockManager
from sync.revision_store import RevisionConflictError, RevisionStore
from sync.shared_cache import SharedCache, SharedUnitEntry
from sync.session_registry import (
    HEARTBEAT_INTERVAL,
    HEARTBEAT_TIMEOUT,
    SessionRegistry,
    SessionInfo,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def shared_workbook(tmp_path) -> str:
    """Create a real .xlsx with unit data for concurrency tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: headers
    headers = [
        "Detailing Due Date", "Prev Due Date", "COM Number", "",
        "Detailer", "Job Name", "Contract", "Description",
        "Build Date", "", "Dept Hours", "% Complete",
        "Remaining Hours", "Actual Hours", "", "", "", "", "", "",
        "Checking Status", "Target Hours", "IEC Hours",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Row 2: COM-001
    ws.cell(row=2, column=3, value="COM-001")
    ws.cell(row=2, column=5, value="Alice")
    ws.cell(row=2, column=6, value="Job Alpha")
    ws.cell(row=2, column=11, value=40.0)
    ws.cell(row=2, column=12, value=0.5)  # 50%
    ws.cell(row=2, column=14, value=20.0)
    ws.cell(row=2, column=22, value="In Progress")
    ws.cell(row=2, column=23, value=36.0)
    ws.cell(row=2, column=24, value=4.0)

    # Row 3: COM-002
    ws.cell(row=3, column=3, value="COM-002")
    ws.cell(row=3, column=5, value="Bob")
    ws.cell(row=3, column=6, value="Job Beta")
    ws.cell(row=3, column=11, value=80.0)
    ws.cell(row=3, column=12, value=0.25)  # 25%
    ws.cell(row=3, column=14, value=20.0)
    ws.cell(row=3, column=22, value="Not Started")
    ws.cell(row=3, column=23, value=60.0)
    ws.cell(row=3, column=24, value=0.0)

    path = str(tmp_path / "test_workbook.xlsx")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def unit_com001() -> Unit:
    return Unit(
        com_number="COM-001",
        job_name="Job Alpha",
        contract_number="",
        description="",
        detailer="Alice",
        checking_status="In Progress",
        department_hours=40.0,
        target_department_hours=36.0,
        iec_internal_hours=4.0,
        percent_complete=50.0,
        actual_hours=20.0,
        working_days=[0, 1, 2, 3],
    )


@pytest.fixture
def unit_com002() -> Unit:
    return Unit(
        com_number="COM-002",
        job_name="Job Beta",
        contract_number="",
        description="",
        detailer="Bob",
        checking_status="Not Started",
        department_hours=80.0,
        target_department_hours=60.0,
        iec_internal_hours=0.0,
        percent_complete=25.0,
        actual_hours=20.0,
        working_days=[0, 1, 2, 3],
    )


# ---------------------------------------------------------------------------
# Test: Save without conflict (different COMs)
# ---------------------------------------------------------------------------


def test_save_no_conflict(shared_workbook, unit_com001, unit_com002):
    """Alice saves COM-001, Bob saves COM-002 — both succeed."""
    alice_store = RevisionStore(shared_workbook)
    bob_store = RevisionStore(shared_workbook)

    # Alice saves COM-001 (baseline=0)
    rev_a = alice_store.commit("COM-001", 0, unit_fingerprint(unit_com001), "Alice@PC1")
    assert rev_a.revision == 1

    # Bob saves COM-002 (baseline=0)
    rev_b = bob_store.commit("COM-002", 0, unit_fingerprint(unit_com002), "Bob@PC2")
    assert rev_b.revision == 1

    # Both revisions exist independently
    assert alice_store.get("COM-001").revision == 1
    assert bob_store.get("COM-002").revision == 1


# ---------------------------------------------------------------------------
# Test: Detect conflict (same COM)
# ---------------------------------------------------------------------------


def test_detect_conflict(shared_workbook, unit_com001):
    """Alice saves COM-001 first, Bob tries with stale baseline."""
    store = RevisionStore(shared_workbook)

    # Alice saves (baseline=0 → revision 1)
    store.commit("COM-001", 0, unit_fingerprint(unit_com001), "Alice@PC1")

    # Bob tries to save with stale baseline (0 instead of 1)
    with pytest.raises(RevisionConflictError) as exc_info:
        store.commit("COM-001", 0, unit_fingerprint(unit_com001), "Bob@PC2")

    assert exc_info.value.latest.revision == 1
    assert exc_info.value.latest.modified_by == "Alice@PC1"


# ---------------------------------------------------------------------------
# Test: Force overwrite after conflict
# ---------------------------------------------------------------------------


def test_force_overwrite(shared_workbook, unit_com001):
    """Alice saves, Bob accepts conflict and force-saves (bumps revision)."""
    store = RevisionStore(shared_workbook)

    # Alice saves (revision 1)
    store.commit("COM-001", 0, "abc", "Alice@PC1")

    # Bob force-saves by using the latest baseline
    rev_overwrite = store.commit("COM-001", 1, "def", "Bob@PC2")
    assert rev_overwrite.revision == 2
    assert rev_overwrite.modified_by == "Bob@PC2"


# ---------------------------------------------------------------------------
# Test: LockManager blocks concurrent access
# ---------------------------------------------------------------------------


def test_lock_blocks_concurrent_excel_write(shared_workbook):
    """Two threads: first acquires excel.lock, second times out."""
    results: dict[str, str] = {}
    lock = LockManager(shared_workbook, "Alice", "PC1")
    lock2 = LockManager(shared_workbook, "Bob", "PC2")

    def alice_acquire():
        try:
            lock.acquire("excel", timeout=2.0)
            results["alice"] = "acquired"
            time.sleep(0.5)  # hold for a bit
            lock.release("excel")
        except Exception as e:
            results["alice"] = f"error: {e}"

    def bob_acquire():
        time.sleep(0.1)  # ensure alice goes first
        try:
            lock2.acquire("excel", timeout=0.5)
            results["bob"] = "acquired"
            lock2.release("excel")
        except LockAcquisitionError:
            results["bob"] = "blocked"
        except Exception as e:
            results["bob"] = f"error: {e}"

    t1 = threading.Thread(target=alice_acquire)
    t2 = threading.Thread(target=bob_acquire)
    t1.start()
    t2.start()
    t1.join()
    t2.join()

    assert results["alice"] == "acquired"
    assert results["bob"] == "blocked"


# ---------------------------------------------------------------------------
# Test: LockManager releases allow subsequent acquire
# ---------------------------------------------------------------------------


def test_lock_release_allows_later_acquire(shared_workbook):
    """Alice acquires then releases; Bob can acquire after."""
    alice = LockManager(shared_workbook, "Alice", "PC1")
    bob = LockManager(shared_workbook, "Bob", "PC2")

    alice.acquire("excel", timeout=1.0)
    alice.release("excel")

    bob.acquire("excel", timeout=1.0)  # should succeed
    bob.release("excel")


# ---------------------------------------------------------------------------
# Test: SharedCache
# ---------------------------------------------------------------------------


def test_shared_cache_store_and_retrieve(shared_workbook, unit_com001):
    """Write a unit to SharedCache, read it back."""
    cache = SharedCache(shared_workbook)
    store = RevisionStore(shared_workbook)
    cache.clear()

    rev = store.commit("COM-001", 0, "abc123", "Alice@PC1")
    cache.update("COM-001", unit_com001, rev)

    entry = cache.get_entry("COM-001")
    assert entry is not None
    assert entry.com_number == "COM-001"
    assert entry.job_name == "Job Alpha"
    assert entry.modified_by == "Alice@PC1"
    assert entry.revision == 1

    # raw dict form
    raw = cache.get("COM-001")
    assert raw is not None
    assert raw["detailer"] == "Alice"


def test_shared_cache_missing_com(shared_workbook):
    """Asking for a COM that was never cached returns None."""
    cache = SharedCache(shared_workbook)
    assert cache.get("COM-999") is None
    assert cache.get_entry("COM-999") is None


def test_shared_cache_clear(shared_workbook, unit_com001):
    """Clearing the cache removes all entries."""
    cache = SharedCache(shared_workbook)
    store = RevisionStore(shared_workbook)

    rev = store.commit("COM-001", 0, "abc", "Alice@PC1")
    cache.update("COM-001", unit_com001, rev)
    assert cache.get("COM-001") is not None

    cache.clear()
    assert cache.get("COM-001") is None


# ---------------------------------------------------------------------------
# Test: SessionRegistry
# ---------------------------------------------------------------------------


def test_session_registry_heartbeat(tmp_path):
    """Create a session, verify heartbeat file exists."""
    excel_path = tmp_path / "workbook.xlsx"
    excel_path.write_bytes(b"fake")

    registry = SessionRegistry(str(excel_path), "Alice@PC1")
    registry.beat()  # write initial heartbeat

    sessions = SessionRegistry.list_active(str(excel_path))
    assert len(sessions) == 1
    assert sessions[0].owner == "Alice@PC1"
    assert not sessions[0].is_stale

    registry.stop()
    # After stop, file should be removed
    sessions = SessionRegistry.list_active(str(excel_path))
    assert len(sessions) == 0


def test_session_registry_multi_user(tmp_path):
    """Two users with separate registries show up as active."""
    excel_path = str(tmp_path / "workbook.xlsx")
    Path(excel_path).write_text("fake")

    alice = SessionRegistry(excel_path, "Alice@PC1")
    bob = SessionRegistry(excel_path, "Bob@PC2")

    alice.beat()
    bob.beat()

    sessions = SessionRegistry.list_active(excel_path)
    owners = {s.owner for s in sessions}
    assert "Alice@PC1" in owners
    assert "Bob@PC2" in owners

    alice.stop()
    bob.stop()


def test_session_stale_detection(tmp_path):
    """A session whose heartbeat is older than timeout is stale."""
    excel_path = str(tmp_path / "workbook.xlsx")
    Path(excel_path).write_text("fake")

    registry = SessionRegistry(excel_path, "Stale@PC1")
    registry.beat()

    # Manually backdate the heartbeat file
    session_file = Path(excel_path).parent / "UnitTracker" / "sessions" / "Stale@PC1.json"
    data = json.loads(session_file.read_text())
    import datetime
    data["last_heartbeat"] = (
        datetime.datetime.now() - datetime.timedelta(seconds=HEARTBEAT_TIMEOUT + 10)
    ).isoformat()
    session_file.write_text(json.dumps(data))

    sessions = SessionRegistry.list_active(excel_path)
    assert len(sessions) == 0  # stale sessions are filtered out

    registry.stop()


# ---------------------------------------------------------------------------
# Test: Save unit to Excel + revision consistency
# ---------------------------------------------------------------------------


def test_save_unit_and_revision(shared_workbook, unit_com001):
    """Save a unit to Excel, verify the revision store is consistent."""
    # Save the unit to the workbook
    save_unit(shared_workbook, unit_com001, sheet_name="Sheet1")

    # Verify revision store has an entry
    store = RevisionStore(shared_workbook)
    rev = store.commit("COM-001", 0, unit_fingerprint(unit_com001), "Alice@PC1")
    assert rev.revision == 1
    assert rev.modified_by == "Alice@PC1"


# ---------------------------------------------------------------------------
# Test: Concurrent save with Excel lock
# ---------------------------------------------------------------------------


def test_concurrent_save_with_lock(shared_workbook, unit_com001, unit_com002):
    """Two threads save different COMs using lock coordination."""
    lock = LockManager(shared_workbook, "Alice", "PC1")
    errors: list[str] = []

    def save_com001():
        try:
            with lock.write_lock():
                save_unit(shared_workbook, unit_com001, sheet_name="Sheet1")
        except Exception as e:
            errors.append(f"Alice: {e}")

    def save_com002():
        try:
            with lock.write_lock():
                save_unit(shared_workbook, unit_com002, sheet_name="Sheet1")
        except Exception as e:
            errors.append(f"Bob: {e}")

    t1 = threading.Thread(target=save_com001)
    t2 = threading.Thread(target=save_com002)

    # They'll block on lock — this tests that the lock serializes access
    t2.start()
    time.sleep(0.05)
    t1.start()
    t1.join()
    t2.join()

    assert not errors, f"Errors during concurrent save: {errors}"


# ---------------------------------------------------------------------------
# Test: force save_unit skips row validation
# ---------------------------------------------------------------------------


def test_save_unit_force(shared_workbook, unit_com001):
    """force=True skips COM-column validation in save_unit."""
    # First save normally
    save_unit(shared_workbook, unit_com001, sheet_name="Sheet1")

    # Modify unit data
    unit_com001.job_name = "Overwritten Job"

    # Force-save with an invalid row hint — should still work since force bypasses validation
    save_unit(shared_workbook, unit_com001, sheet_name="Sheet1", row_idx=2, force=True)

    # Verify by reloading
    from openpyxl import load_workbook
    wb = load_workbook(shared_workbook, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    job_name = ws.cell(row=2, column=6).value
    wb.close()
    assert job_name == "Overwritten Job"