#!/usr/bin/env python3
"""
Spike 001: Create SQLite database and migrate data from the Excel workbook.

Reads the 'Current List' sheet from the existing workbook and migrates all
rows into a SQLite database with proper types and indexes.

Also seeds the detailers table and default_schedule from config.yaml data.

Usage:
    python3 migrate_workbook_to_sqlite.py [--workbook PATH] [--db PATH]
"""
import argparse
import json
import logging
import os
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS units (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    com_number TEXT UNIQUE NOT NULL,

    -- RAW IMPORT (from SSRS CSV)
    detailing_due_date TEXT,
    manufacturing_location TEXT,
    job_name TEXT,
    top_level_number TEXT,
    description TEXT,
    build_date TEXT,
    build_cycle INTEGER,
    department_hours REAL DEFAULT 0.0,
    week_ending_friday TEXT,

    -- PIPELINE-DERIVED
    dept_due_date_previous TEXT,
    remaining_hours REAL DEFAULT 0.0,

    -- MANUAL/USER
    detailer TEXT,
    percent_complete REAL DEFAULT 0.0,
    actual_hours REAL DEFAULT 0.0,
    notes TEXT,
    late INTEGER DEFAULT 0,
    checking_status TEXT,
    target_dept_hours REAL DEFAULT 0.0,
    iec_internal_hours REAL DEFAULT 0.0,
    unit_detailing_start_date TEXT,
    unit_moved_to_checking_date TEXT,
    unit_detailing_completion_date TEXT,
    actual_hours_to_detail_unit REAL DEFAULT 0.0,
    hour_variance REAL DEFAULT 0.0,
    remaining_demand REAL DEFAULT 0.0,
    same_as TEXT,
    dr_checks TEXT,
    dvl_checks TEXT,
    hours_checking REAL DEFAULT 0.0,

    -- COMPUTED (precomputed in Python, stored for fast reads)
    working_days_until_due INTEGER,
    calendar_days_until_due INTEGER,
    days_diff_due_to_build INTEGER,

    -- Metadata
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now'))
);

-- Detailers and their working schedules
CREATE TABLE IF NOT EXISTS detailers (
    name TEXT PRIMARY KEY,
    working_weekdays TEXT NOT NULL,
    display_order INTEGER DEFAULT 0
);

-- Default schedule for unassigned/new detailers
CREATE TABLE IF NOT EXISTS default_schedule (
    id INTEGER PRIMARY KEY CHECK (id = 1),
    working_weekdays TEXT NOT NULL
);

-- Indexes
CREATE INDEX IF NOT EXISTS idx_units_com ON units(com_number);
CREATE INDEX IF NOT EXISTS idx_units_build_date ON units(build_date);
CREATE INDEX IF NOT EXISTS idx_units_week_ending ON units(week_ending_friday);
CREATE INDEX IF NOT EXISTS idx_units_detailer ON units(detailer);
CREATE INDEX IF NOT EXISTS idx_units_pct_complete ON units(percent_complete);
CREATE INDEX IF NOT EXISTS idx_units_manufacturing_location ON units(manufacturing_location);
CREATE INDEX IF NOT EXISTS idx_units_working_days ON units(working_days_until_due);
CREATE INDEX IF NOT EXISTS idx_units_due_date ON units(detailing_due_date);
"""

SEED_SQL = """
INSERT OR IGNORE INTO default_schedule (id, working_weekdays) VALUES (1, '[0,1,2,3]');

INSERT OR IGNORE INTO detailers (name, working_weekdays, display_order) VALUES
    ('Jackie H',  '[0,1,2,3]', 1),
    ('Tommy N',   '[1,2,3,4]', 2),
    ('Matthew S', '[1,2,3,4]', 3),
    ('Matthew E', '[1,2,3,4]', 4),
    ('Carl M',    '[0,1,2,3]', 5),
    ('Stewart D', '[1,2,3,4]', 6),
    ('David K',   '[0,1,2,3]', 7),
    ('Katie D',   '[0,1,2,3]', 8),
    ('Kris L',    '[0,1,2,3]', 9),
    ('Emilio P',  '[0,1,2,3]', 10),
    ('Timothy B', '[1,2,3,4]', 11),
    ('Jeremy B',  '[0,1,2,3]', 12),
    ('Brandon B', '[1,2,3,4]', 13),
    ('Tracy V',   '[0,1,2,3]', 14),
    ('Tanner D',  '[0,1,2,3]', 15);
"""

# ---------------------------------------------------------------------------
# Current List sheet → SQLite column mapping
# ---------------------------------------------------------------------------

COLUMN_MAP = {
    "A": ("detailing_due_date", "date"),
    "B": ("dept_due_date_previous", "str"),  # store as-is (legacy noise + dates)
    "C": ("com_number", "str"),
    "D": ("manufacturing_location", "str"),
    "E": ("detailer", "str"),
    "F": ("job_name", "str"),
    "G": ("top_level_number", "str"),
    "H": ("description", "str"),
    "I": ("build_date", "date"),
    "J": ("build_cycle", "int"),
    "K": ("department_hours", "float"),
    "L": ("percent_complete", "percent"),
    "M": ("remaining_hours", "float"),
    "N": ("actual_hours", "float"),
    "O": ("week_ending_friday", "date"),
    "P": ("notes", "str"),
    "Q": ("late", "int"),
    # R, S, T are COMPUTED — not stored in units table
    "U": ("checking_status", "str"),
    "V": ("target_dept_hours", "float"),
    "W": ("iec_internal_hours", "float"),
    "X": ("unit_detailing_start_date", "date"),
    "Y": ("unit_moved_to_checking_date", "date"),
    "Z": ("unit_detailing_completion_date", "date"),
    "AA": ("actual_hours_to_detail_unit", "float"),
    "AB": ("hour_variance", "float"),
    "AC": ("remaining_demand", "float"),
    "AD": ("same_as", "str"),
    "AE": ("dr_checks", "str"),
    "AF": ("dvl_checks", "str"),
    "AG": ("hours_checking", "float"),
    "AK": ("days_diff_due_to_build", "int"),
}

# ---------------------------------------------------------------------------
# Type converters
# ---------------------------------------------------------------------------

def convert_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        if len(val) >= 10 and val[4] == '-':
            return val[:10]
        log.warning(f"Could not parse date: {val!r}")
        return None
    if isinstance(val, (int, float)):
        try:
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=int(val))).strftime("%Y-%m-%d")
        except (OverflowError, ValueError, OSError):
            pass
    return None


def convert_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).strip().replace(",", "").replace("$", "")
    if not val:
        return None
    try:
        return float(val)
    except ValueError:
        return None


def convert_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val = str(val).strip()
    if not val:
        return None
    try:
        return int(float(val))
    except ValueError:
        return None


def convert_percent(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        if val == 0:
            return 0.0
        if val == 1:
            return 1.0
        return None
    if isinstance(val, float):
        if val > 1.0:
            return None
        return val
    val = str(val).strip().replace("%", "")
    if not val:
        return None
    try:
        fval = float(val)
        if fval > 1.0:
            return fval / 100.0
        return fval
    except ValueError:
        return None


def convert_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


CONVERTERS = {
    "date": convert_date,
    "float": convert_float,
    "int": convert_int,
    "percent": convert_percent,
    "str": convert_str,
}

# ---------------------------------------------------------------------------
# working_days_until_due computation
# ---------------------------------------------------------------------------

def working_days_between(start: date, end: date, working_weekdays: list[int]) -> int:
    count = 0
    current = start + timedelta(days=1)
    while current <= end:
        if current.weekday() in working_weekdays:
            count += 1
        current += timedelta(days=1)
    return count


def get_working_weekdays(cursor, detailer_name):
    if detailer_name:
        cursor.execute("SELECT working_weekdays FROM detailers WHERE name = ?", (detailer_name,))
        row = cursor.fetchone()
        if row:
            return json.loads(row[0])
    cursor.execute("SELECT working_weekdays FROM default_schedule WHERE id = 1")
    row = cursor.fetchone()
    if row:
        return json.loads(row[0])
    return [0, 1, 2, 3]


def recompute_due_columns_batch(cursor):
    """Recompute working_days_until_due and calendar_days_until_due for ALL rows."""
    today = date.today()
    # Get all detailers' schedules in one query
    cursor.execute("SELECT name, working_weekdays FROM detailers")
    detailer_map = {r[0]: json.loads(r[1]) for r in cursor.fetchall()}
    cursor.execute("SELECT working_weekdays FROM default_schedule WHERE id = 1")
    default_row = cursor.fetchone()
    default_weekdays = json.loads(default_row[0]) if default_row else [0, 1, 2, 3]

    cursor.execute(
        "SELECT com_number, detailing_due_date, detailer FROM units WHERE detailing_due_date IS NOT NULL"
    )
    rows = cursor.fetchall()

    for com, due_str, detailer in rows:
        try:
            due_date = date.fromisoformat(due_str)
        except ValueError:
            continue
        calendar_days = (due_date - today).days
        weekdays = detailer_map.get(detailer, default_weekdays) if detailer else default_weekdays
        working_days = working_days_between(today, due_date, weekdays) if calendar_days >= 0 else 0
        cursor.execute(
            "UPDATE units SET calendar_days_until_due = ?, working_days_until_due = ? WHERE com_number = ?",
            (calendar_days, working_days, com)
        )


# ---------------------------------------------------------------------------
# Migration
# ---------------------------------------------------------------------------

def col_letter_to_index(letter: str) -> int:
    idx = 0
    for ch in letter.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


INSERT_FIELDS = []
for col_letter in sorted(COLUMN_MAP.keys(), key=lambda x: col_letter_to_index(x)):
    field_name, converter_type = COLUMN_MAP[col_letter]
    INSERT_FIELDS.append((col_letter, field_name, converter_type))


def create_database(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.executescript(SCHEMA_SQL)
    conn.executescript(SEED_SQL)
    conn.commit()
    log.info(f"Created database at {db_path}")
    return conn


def migrate_workbook(workbook_path: str, db_path: str, sheet_name: str = "Current List"):
    from openpyxl import load_workbook

    log.info(f"Loading workbook: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
    except KeyError:
        log.error(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    conn = create_database(db_path)
    cur = conn.cursor()

    field_col_indices = []
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[cell.column_letter] = cell.column - 1

    log.info(f"Header columns found: {len(header_map)}")

    for col_letter, field_name, converter_type in INSERT_FIELDS:
        wb_col_idx = header_map.get(col_letter)
        field_col_indices.append((field_name, CONVERTERS[converter_type], wb_col_idx))

    fields_str = ", ".join(fn for _, fn, _ in INSERT_FIELDS)
    placeholders_str = ", ".join(["?"] * len(INSERT_FIELDS))
    updates_str = ", ".join(
        f"{fn} = excluded.{fn}" for _, fn, _ in INSERT_FIELDS if fn != "com_number"
    )

    insert_sql = f"""
        INSERT INTO units ({fields_str})
        VALUES ({placeholders_str})
        ON CONFLICT(com_number) DO UPDATE SET
            {updates_str}
    """

    total_rows = 0
    inserted = 0
    updated = 0
    errors = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        total_rows += 1

        com_idx = header_map.get("C")
        if com_idx is None:
            log.error("Could not find COM number column (C)")
            sys.exit(1)
        com_number = row[com_idx] if com_idx < len(row) else None
        if com_number is None or str(com_number).strip() == "":
            continue
        com_number = str(com_number).strip()

        values = []
        all_none = True
        for field_name, converter, wb_col_idx in field_col_indices:
            raw = row[wb_col_idx] if wb_col_idx is not None and wb_col_idx < len(row) else None
            converted = converter(raw)
            values.append(converted)
            if converted is not None:
                all_none = False

        if all_none:
            continue

        try:
            cur.execute(insert_sql, values)
            if cur.rowcount == 1:
                if cur.lastrowid:
                    inserted += 1
                else:
                    updated += 1
            conn.commit()
        except Exception as e:
            errors += 1
            log.error(f"Row {row_idx} (COM {com_number}): {e}")
            if errors <= 5:
                log.error(f"  Values: {values[:5]}...")
            conn.rollback()
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()

        if total_rows % 5000 == 0:
            log.info(f"  Processed {total_rows} rows...")

    wb.close()

    # Recompute due columns for all migrated rows
    log.info("Recomputing working_days_until_due and calendar_days_until_due...")
    recompute_due_columns_batch(cur)
    conn.commit()

    # Verification
    cur.execute("SELECT COUNT(*) FROM units")
    db_count = cur.fetchone()[0]
    cur.execute("SELECT SUM(department_hours) FROM units")
    db_hours = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM detailers")
    detailer_count = cur.fetchone()[0]

    log.info("")
    log.info("=" * 60)
    log.info("MIGRATION COMPLETE")
    log.info("=" * 60)
    log.info(f"Workbook rows read:    {total_rows}")
    log.info(f"Rows inserted:          {inserted}")
    log.info(f"Rows updated:           {updated}")
    log.info(f"Errors:                 {errors}")
    log.info(f"DB row count:           {db_count}")
    log.info(f"DB SUM(dept_hrs):       {db_hours:.2f}")
    log.info(f"Detailers seeded:       {detailer_count}")
    log.info("=" * 60)

    conn.close()


def main():
    parser = argparse.ArgumentParser(description="Migrate Excel workbook to SQLite")
    parser.add_argument("--workbook", default=None, help="Path to workbook (.xlsm)")
    parser.add_argument("--db", default=None, help="Output SQLite database path")
    parser.add_argument("--sheet", default="Current List", help="Sheet name to read")
    args = parser.parse_args()

    workbook = args.workbook or str(Path(__file__).resolve().parent.parent /
        "SQL-Schedule-App" /
        "SCHDetailingReport_all_plants_MASTER")
    db = args.db or str(Path(__file__).resolve().parent.parent /
        "SQL-Schedule-App" / "schedule.db")
    workbook = str(Path(workbook).resolve())
    db = str(Path(db).resolve())

    log.info(f"Workbook: {workbook}")
    log.info(f"Database: {db}")

    if not os.path.exists(workbook):
        log.error(f"Workbook not found: {workbook}")
        sys.exit(1)

    migrate_workbook(workbook, db, args.sheet)


if __name__ == "__main__":
    main()